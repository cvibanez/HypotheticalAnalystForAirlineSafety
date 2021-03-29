# Conrad Ibanez
# DSC640 Winter 2020
# Project Task 4: Infographic

import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
import openpyxl


def read_airline_safety_file():
    # read the csv file
    airline_safety_df = pd.read_csv("../../data/airline-safety.csv")
    # Remove asterisk in airline name
    airline_safety_df['airline'] = airline_safety_df['airline'].str.replace('*','')
    return airline_safety_df


def read_transportation_expenditures_file():
    # get data
    file = "../../data/task3/PersonalConsumptionExpenditures_USDeptCommerce.xlsx"
    
    # Load workbook
    wb = openpyxl.load_workbook(file)
    #Access to a worksheet
    ws = wb['Sheet1']
    
    data= ws.values
    
    # Get the first line in file as a header line
    columns = next(data)[0:]
 
    # Convert to DataFrame
    data_df = pd.DataFrame(data, columns=columns)
    
    return data_df

def read_passengers_file():
    # get data
    file = "../../data/Task4/System_Total_Enplaned_Passengers_Clean.xlsx"

    # Load workbook
    wb = openpyxl.load_workbook(file)
    #Access to a worksheet
    ws = wb['Sheet1']
    
    data= ws.values
    
    # Get the first line in file as a header line
    columns = next(data)[0:]
 
    # Convert to DataFrame
    data_df = pd.DataFrame(data, columns=columns)
    
    return data_df

def read_airline_accidents_file():

    # get data
    file = "../../data/Task4/table9_2014-accidents_ntsb-classification-1995-2014-for-u-s-air-carriers_Clean.xlsx"

    # Load workbook
    wb = openpyxl.load_workbook(file)
    #Access to a worksheet
    ws = wb['table9_2014-accidents_ntsb-clas']
    
    data= ws.values
    
    # Get the first line in file as a header line
    columns = next(data)[0:]
 
    # Convert to DataFrame
    data_df = pd.DataFrame(data, columns=columns)
    
    return data_df

def read_auto_accidents_file():
    # get data
    file = "../../data/Task4/National Statistics_Clean.xlsx"

    # Load workbook
    wb = openpyxl.load_workbook(file)
    #Access to a worksheet
    ws = wb['Sheet1']
    
    data= ws.values
    
    # Get the first line in file as a header line
    columns = next(data)[0:]
 
    # Convert to DataFrame
    data_df = pd.DataFrame(data, columns=columns)
    
    return data_df

def read_time_performance_file():
    # get data
    file = "../../data/Task4/Table1_45_Summary_Airline_TimePerformance_ThruJan2017_Clean.xlsx"

    # Load workbook
    wb = openpyxl.load_workbook(file)
    #Access to a worksheet
    ws = wb['Table1']
    
    data= ws.values
    
    # Get the first line in file as a header line
    columns = next(data)[0:]
 
    # Convert to DataFrame
    data_df = pd.DataFrame(data, columns=columns)
    
    return data_df


# Reference- https://plotly.com/python/horizontal-bar-charts/
def create_airline_safety_horizontal_bar():
    # get data
    airline_safety_df = read_airline_safety_file()
    us_airlines = ['United / Continental',
                   'Delta / Northwest',
                   'American',
                   'Southwest Airlines']
    
    # Sort data by available seat km
    airline_safety_df = airline_safety_df.sort_values(by=['avail_seat_km_per_week'])
    airline_safety_df = airline_safety_df[airline_safety_df['airline'].isin(us_airlines)]
    
    airline_names = airline_safety_df['airline']
    
    airlines = airline_names.tolist()
    y_pos = np.arange(len(airlines))
    quantity = airline_safety_df['fatalities_00_14'].tolist()
    
    fig, ax = plt.subplots(figsize=(3,2))

    ax.barh(y_pos, quantity)
    ax.set_yticks(y_pos)
    ax.set_yticklabels(airlines)
    #ax.invert_yaxis()  # labels read top-to-bottom
    ax.set_xlabel('Fatalities')
    #ax.set_ylabel('Airline')  # Looks better without as title also shows
    ax.set_title('Top U.S. Airlines 2000-2014 Fatalities')

    plt.show()
    
    # https://datatofish.com/line-chart-python-matplotlib/
def create_expenditures_airline():
    # get data
    data_df = read_transportation_expenditures_file()
    print(data_df.head())
    
    # Reference https://stackoverflow.com/questions/58812886/pandas-line-plot-without-transposing-the-dataframe
    x = data_df.columns.tolist()[1:]
    y = data_df.iloc[3:4, 1:].values
    print("x:", x)
    print("y:", y)
    for i, j in enumerate(y):
        plt.plot(x, j, label=data_df['Transportation'].iloc[3])
    #plt.ylim(bottom=0)
    plt.legend(bbox_to_anchor=(1.01, 1), loc='upper left') # 
    plt.xlabel('Year')
    plt.ylabel('U.S. dollars (millions)')  # Looks better without as title also shows
    plt.title('Consumer Airline Expenditures 2010-2018')
    plt.show()
    
# https://datatofish.com/line-chart-python-matplotlib/
def create_passengers_line_chart():
    # get data
    airline_df =  read_passengers_file()
    airline_array = ['United','Delta','American','Southwest']
    # Need to transpose the data due to the file setup
    #https://towardsdatascience.com/transform-reality-with-pandas-96f061628030
    airline_df =  airline_df.set_index('Airline')
    airline_df = airline_df.transpose()
    #print(airline_df)  
    #print(airline_df['American'])
      
    plt.figure(figsize=(6, 4)) 
    years_list = airline_df.index.values.flatten().tolist()
    years = np.arange(len(years_list))
    
    for airline in airline_array:
        plt.plot(years, airline_df[airline], label=airline)
        print(airline, "- enplaned- ", airline_df[airline] )
        
    plt.title('U.S. Airline Enplaned Passengers')
    plt.xlabel('Year')
    plt.ylabel('Passengers (thousands)')
    plt.xticks(years, years_list, rotation='vertical', fontsize = 10)
    plt.legend(bbox_to_anchor=(1.01, 1), loc='upper left')
    plt.show()

# Reference- https://plotly.com/python/bar-charts/    
def create_accidents_stacked_bar():
    # get data
    accidents = read_airline_accidents_file()
    
    accidents['Accidents_Non_Fatal'] = accidents['Accidents_All'] - accidents['Accidents_Fatal']
    
    # Convert Year values to numeric to be able to filter
    accidents[['Year']] = accidents[['Year']].apply(pd.to_numeric)
    # Only get years less than or equal to 2014
    accidents = accidents[(accidents['Year'] >= 2010) & (accidents['Year'] <= 2014)]
    
    #years_list = accidents['Year']values.astype(int).flatten().tolist()
    years_list = accidents['Year'].values.flatten().tolist()
#    first_qty =  accidents['Accidents_Fatal'].values.flatten().values.flatten().tolist()
#    second_qty =  accidents['Accident _Non_Fatal'].values.flatten().values.flatten().tolist()
    
    first_qty =  accidents['Accidents_Fatal']
    second_qty =  accidents['Accidents_Non_Fatal']

    
#    # Heights of bars1 + bars2
#    bars_first_second = np.add(first_qty, second_qty).tolist()

    
    years = np.arange(len(years_list))    
    width = 0.8      # the width of the bars: can also be len(x) sequence
    
    plt.figure(figsize=(4, 3)) 
    
    p1 = plt.bar(years, first_qty, width)
    p2 = plt.bar(years, second_qty,  width, bottom=first_qty)
    plt.ylabel('Accidents')
    plt.title('U.S. Airline Accidents from 2010-2014')
    plt.xticks(years, years_list, rotation='vertical', fontsize = 10)
    #plt.yticks(np.arange(0, 81, 10))
    plt.legend((p1[0], p2[0]), ('Fatal', 'Non-Fatal'), loc="upper right", ncol=2, fontsize='small')
    plt.show()
    
# Reference- https://plotly.com/python/bar-charts/    
def create_auto_accidents_stacked_bar():
    # get data
    accidents = read_auto_accidents_file()
    
    # Convert Year values to numeric to be able to filter
    accidents[['Year']] = accidents[['Year']].apply(pd.to_numeric)
    # Only get years less than or equal to 2014
    accidents = accidents[accidents['Year'] <= 2014]
    
    #accidents['Accidents_Non_Fatal'] = accidents['Accidents_All'] - accidents['Accidents_Fatal']
    
    #years_list = accidents['Year']values.astype(int).flatten().tolist()
    years_list = accidents['Year'].values.flatten().tolist()
#    first_qty =  accidents['Accidents_Fatal'].values.flatten().values.flatten().tolist()
#    second_qty =  accidents['Accident _Non_Fatal'].values.flatten().values.flatten().tolist()
    
    first_qty =  accidents['Fatal']
    second_qty =  accidents['Injury']
    third_qty =  accidents['Property-Damage-Only']


    
#    # Heights of bars1 + bars2
    bars_first_second = np.add(first_qty, second_qty).tolist()

    
    years = np.arange(len(years_list))    
    width = 0.8      # the width of the bars: can also be len(x) sequence
    
    plt.figure(figsize=(6, 10)) 
    
    p1 = plt.bar(years, first_qty, width)
    p2 = plt.bar(years, second_qty,  width, bottom=first_qty)
    p3 = plt.bar(years, third_qty,  width, bottom=bars_first_second)
    plt.ylabel('Accidents')
    plt.title('U.S. Auto Accidents from 2010-2014')
    plt.xticks(years, years_list, rotation='vertical', fontsize = 10)
    #plt.yticks(useOffset=False)
    plt.ticklabel_format(axis='y', style='plain')
    plt.legend((p1[0], p2[0], p3[0]), ('Fatal', 'Injury', 'Property Damage Only'), loc="upper left", ncol=3, fontsize='small')
    plt.show()

def create_time_performance_graph():
    # get data
    data_df = read_time_performance_file()
    
    # Reference https://stackoverflow.com/questions/58812886/pandas-line-plot-without-transposing-the-dataframe
    plt.plot(data_df['Year'], data_df['Percent_On-Time_Arrivals'])
    
    # Plot the mean reference https://py.plainenglish.io/line-chart-basics-with-pythons-matplotlib-e52032981bd3
    arrival_mean = data_df['Percent_On-Time_Arrivals'].mean()
    print(arrival_mean)
    mean_array = np.ones(data_df.shape[0])*arrival_mean
    
    plt.plot(data_df['Year'], mean_array, color='black', linewidth=2, linestyle='--')
    #plt.ylim(bottom=0)
    plt.legend(bbox_to_anchor=(1.01, 1), loc='upper left') # 
    plt.xlabel('Year')
    plt.ylabel('On Time Arrival Percentage')
    plt.title('U.S. Airlines On Time Arrival Percentage')
    plt.show()
    

def main():
    
    #create_airline_safety_horizontal_bar()
    #create_expenditures_airline()
    create_passengers_line_chart()
 
    #create_accidents_stacked_bar()
    #create_auto_accidents_stacked_bar()
    #create_time_performance_graph()
    
if __name__ == '__main__':
    main()
    
